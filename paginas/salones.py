"""
Página de salones: ocupación, disponibilidad y uso.
Con detección correcta de choques y descarga a Excel.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.utils.queries import (
    buscar_salones, clases_en_salon, cargar_periodos, cargar_tipos_salon
)


DIAS_ORDEN = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']
DIAS_CORTO = {
    'LUNES': 'LUN', 'MARTES': 'MAR', 'MIERCOLES': 'MIÉ',
    'JUEVES': 'JUE', 'VIERNES': 'VIE', 'SABADO': 'SÁB', 'DOMINGO': 'DOM'
}


def hora_a_minutos(hora_str):
    partes = str(hora_str).split(':')
    return int(partes[0]) * 60 + int(partes[1])


def minutos_a_hora(minutos):
    return f"{minutos // 60:02d}:{minutos % 60:02d}"


def clases_se_solapan(h1, h2):
    """Devuelve True solo si dos clases REALMENTE se solapan en tiempo."""
    ini1 = hora_a_minutos(h1['hora_inicio'])
    fin1 = hora_a_minutos(h1['hora_fin'])
    ini2 = hora_a_minutos(h2['hora_inicio'])
    fin2 = hora_a_minutos(h2['hora_fin'])
    return ini1 < fin2 and ini2 < fin1

def construir_horario_cuadricula(horarios):
    """
    Construye horario tipo cuadrícula que se adapta:
    - Si todos los horarios son en horas en punto → bloques de 1 hora
    - Si hay horarios "rotos" (X:30, X:29) → bloques de 30 minutos
    """
    if not horarios:
        return None, None
    
    # 1. Detectar si hay horarios que NO empiecen o terminen en hora en punto
    necesita_media_hora = False
    for h in horarios:
        ini_min = hora_a_minutos(h['hora_inicio'])
        fin_min = hora_a_minutos(h['hora_fin'])
        
        # Si el inicio NO está en una hora en punto (no es múltiplo de 60)
        # O si el fin no es "casi" hora en punto (terminar en :59 sí es válido)
        if ini_min % 60 != 0 or (fin_min + 1) % 60 != 0:
            necesita_media_hora = True
            break
    
    # 2. Definir el tamaño del bloque (60 minutos o 30 minutos)
    tamano_bloque = 30 if necesita_media_hora else 60
    
    # 3. Calcular el rango total del horario
    hora_min_minutos = min(hora_a_minutos(h['hora_inicio']) for h in horarios)
    hora_max_minutos = max(hora_a_minutos(h['hora_fin']) for h in horarios)
    
    # Redondear al múltiplo del tamaño del bloque
    hora_inicio_grid = (hora_min_minutos // tamano_bloque) * tamano_bloque
    hora_fin_grid = ((hora_max_minutos + tamano_bloque - 1) // tamano_bloque) * tamano_bloque
    
    # 4. Construir filas
    filas_df = []
    info_choques = []
    
    for bloque_actual in range(hora_inicio_grid, hora_fin_grid, tamano_bloque):
        bloque_inicio_str = minutos_a_hora(bloque_actual)
        bloque_fin_str = minutos_a_hora(bloque_actual + tamano_bloque - 1)
        rango_str = f"{bloque_inicio_str} - {bloque_fin_str}"
        
        fila = {"HORA": rango_str}
        
        for dia in DIAS_ORDEN:
            dia_corto = DIAS_CORTO[dia]
            
            clases_en_bloque = []
            for h in horarios:
                if h['dia_semana'] != dia:
                    continue
                
                clase_ini = hora_a_minutos(h['hora_inicio'])
                clase_fin = hora_a_minutos(h['hora_fin'])
                
                # ¿La clase se solapa con este bloque?
                if clase_fin > bloque_actual and clase_ini < bloque_actual + tamano_bloque:
                    clases_en_bloque.append(h)
            
            if not clases_en_bloque:
                fila[dia_corto] = "—"
                continue
            
            # Detectar si REALMENTE se solapan entre sí
            hay_solapamiento = False
            if len(clases_en_bloque) > 1:
                for i in range(len(clases_en_bloque)):
                    for j in range(i+1, len(clases_en_bloque)):
                        if clases_se_solapan(clases_en_bloque[i], clases_en_bloque[j]):
                            hay_solapamiento = True
                            info_choques.append({
                                "dia": dia,
                                "clase_1": clases_en_bloque[i],
                                "clase_2": clases_en_bloque[j]
                            })
                            break
                    if hay_solapamiento:
                        break
            
            # Construir texto de la celda
            textos = []
            for c in clases_en_bloque:
                clase_info = c.get('clases') or {}
                materia = (clase_info.get('materias') or {})
                nombre = materia.get('descripcion') or '(multi)'
                crn = c['crn']
                textos.append(f"{nombre[:25]} (CRN {crn})")
            
            if hay_solapamiento:
                fila[dia_corto] = "⚠️ " + " ║ ".join(textos)
            else:
                fila[dia_corto] = " · ".join(textos) if len(textos) > 1 else textos[0]
        
        filas_df.append(fila)
    
    df = pd.DataFrame(filas_df)
    return df, info_choques

def generar_excel_salon(salon_obj, horarios, df_detalle, df_cuadricula):
    """Genera un archivo Excel con formato bonito del horario del salón."""
    wb = Workbook()
    
    # ===== HOJA 1: Resumen =====
    ws1 = wb.active
    ws1.title = "Resumen"
    
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
    bold_font = Font(bold=True)
    
    # Encabezado
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"HORARIO DEL SALÓN {salon_obj['codigo']}"
    ws1['A1'].font = titulo_font
    ws1['A1'].fill = titulo_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 25
    
    # Datos del salón
    ws1['A3'] = "Código:"
    ws1['A3'].font = bold_font
    ws1['B3'] = salon_obj['codigo']
    
    ws1['A4'] = "Descripción:"
    ws1['A4'].font = bold_font
    ws1['B4'] = salon_obj.get('descripcion', 'N/A')
    
    ws1['A5'] = "Capacidad:"
    ws1['A5'].font = bold_font
    ws1['B5'] = salon_obj.get('capacidad', 0)
    
    ws1['A6'] = "Tipo:"
    ws1['A6'].font = bold_font
    ws1['B6'] = salon_obj.get('tipo_uso_descripcion', 'N/A')
    
    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 40
    
    # ===== HOJA 2: Horario tradicional (cuadrícula) =====
    ws2 = wb.create_sheet("Horario semanal")
    
    if df_cuadricula is not None and not df_cuadricula.empty:
        # Encabezado
        ws2.merge_cells('A1:H1')
        ws2['A1'] = f"HORARIO SEMANAL · {salon_obj['codigo']}"
        ws2['A1'].font = titulo_font
        ws2['A1'].fill = titulo_fill
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 25
        
        # Headers
        headers = ['HORA', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
        header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Datos
        ocupada_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        libre_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        choque_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        for row_idx, row in enumerate(df_cuadricula.values, 4):
            for col_idx, valor in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=str(valor))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                cell.font = Font(size=9)
                
                if col_idx == 1:  # Columna HORA
                    cell.font = Font(bold=True, size=10)
                    cell.fill = header_fill
                elif valor == "—":
                    cell.fill = libre_fill
                    cell.font = Font(size=10, color="9E9E9E")
                elif "⚠️" in str(valor):
                    cell.fill = choque_fill
                else:
                    cell.fill = ocupada_fill
            
            ws2.row_dimensions[row_idx].height = 35
        
        # Ancho de columnas
        ws2.column_dimensions['A'].width = 15
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws2.column_dimensions[col_letter].width = 22
    
    # ===== HOJA 3: Detalle listado =====
    ws3 = wb.create_sheet("Detalle")
    
    if df_detalle is not None and not df_detalle.empty:
        # Encabezados
        for col_idx, col_name in enumerate(df_detalle.columns, 1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = bold_font
            cell.fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, row in enumerate(df_detalle.values, 2):
            for col_idx, valor in enumerate(row, 1):
                ws3.cell(row=row_idx, column=col_idx, value=str(valor))
        
        # Auto-ajustar anchos
        for col_idx in range(1, len(df_detalle.columns) + 1):
            max_len = max(len(str(df_detalle.iloc[i, col_idx-1])) for i in range(len(df_detalle)))
            max_len = max(max_len, len(df_detalle.columns[col_idx-1]))
            ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    
    # Guardar a bytes para descargar
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def main():
    st.title("🚪 Salones")
    st.markdown("Información de salones físicos y su ocupación")
    st.divider()
    
    col1, col2, col3 = st.columns([2, 2, 1])
    with col1:
        codigo_busqueda = st.text_input("🔍 Buscar por código", placeholder="Ej: 11-A001")
    with col2:
        tipos = cargar_tipos_salon()
        tipo_filtro = st.selectbox("🏷️ Tipo de salón", ["Todos"] + tipos)
    with col3:
        st.write("")
        st.write("")
        buscar = st.button("🔍 Buscar", type="primary", use_container_width=True)
    
    salones = buscar_salones(codigo_busqueda, tipo_filtro)
    
    if not salones:
        st.warning("⚠️ No se encontraron salones con esos filtros")
        return
    
    st.caption(f"📊 {len(salones)} salones encontrados")
    
    opciones = [f"{s['codigo']} · {s.get('descripcion', '')} · Capacidad: {s.get('capacidad', 0)}" for s in salones]
    seleccion = st.selectbox("Selecciona un salón para ver su ocupación", opciones)
    
    salon_codigo = seleccion.split(" · ")[0]
    salon_obj = next(s for s in salones if s['codigo'] == salon_codigo)
    
    st.divider()
    
    st.header(f"🚪 {salon_obj['codigo']}")
    
    col_m1, col_m2, col_m3 = st.columns(3)
    with col_m1:
        st.metric("📝 Descripción", salon_obj.get('descripcion', 'N/A'))
    with col_m2:
        st.metric("👥 Capacidad", salon_obj.get('capacidad', 0))
    with col_m3:
        st.metric("🏷️ Tipo", salon_obj.get('tipo_uso_descripcion', 'N/A'))
    
    periodos = cargar_periodos()
    opciones_periodo = ["Todos"] + [f"{p['id']}" for p in periodos]
    periodo_sel = st.selectbox("Filtrar por periodo", opciones_periodo, key="periodo_salon")
    periodo_id = int(periodo_sel) if periodo_sel != "Todos" else None
    
    horarios = clases_en_salon(salon_codigo, periodo_id)
    
    if not horarios:
        st.info("Este salón no tiene clases asignadas en el filtro seleccionado")
        return
    
    # Métricas
    horas_uso = 0
    for h in horarios:
        try:
            ini = hora_a_minutos(h['hora_inicio'])
            fin = hora_a_minutos(h['hora_fin'])
            horas_uso += (fin - ini) / 60
        except Exception:
            pass
    
    horas_disponibles = 90
    porcentaje_uso = (horas_uso / horas_disponibles * 100) if horas_disponibles > 0 else 0
    
    col_u1, col_u2, col_u3 = st.columns(3)
    with col_u1:
        st.metric("📊 Clases asignadas", len(set((h['crn'], h['periodo_id']) for h in horarios)))
    with col_u2:
        st.metric("⏰ Horas/semana", f"{horas_uso:.1f}")
    with col_u3:
        st.metric("📈 % de uso", f"{porcentaje_uso:.1f}%")
    
    st.progress(min(porcentaje_uso / 100, 1.0), text=f"Ocupación semanal: {porcentaje_uso:.1f}% (de 90 hrs/semana disponibles)")
    
    st.divider()
    
    # Detalle lineal
    st.subheader("📋 Ocupación detallada")
    
    horarios_ordenados = sorted(
        horarios,
        key=lambda h: (DIAS_ORDEN.index(h['dia_semana']) if h['dia_semana'] in DIAS_ORDEN else 99, h['hora_inicio'])
    )
    
    filas = []
    for h in horarios_ordenados:
        clase_info = h.get('clases') or {}
        materia = (clase_info.get('materias') or {})
        maestro = (clase_info.get('maestros') or {})
        
        filas.append({
            "Día": h['dia_semana'],
            "Hora": f"{h['hora_inicio'][:5]} - {h['hora_fin'][:5]}",
            "CRN": h['crn'],
            "Periodo": h['periodo_id'],
            "Grupo": clase_info.get('grupo') or '',
            "Materia": materia.get('descripcion') or '(multi)',
            "Maestro": maestro.get('nombre_completo') or 'Sin asignar'
        })
    
    df_detalle = pd.DataFrame(filas)
    st.dataframe(df_detalle, use_container_width=True, hide_index=True, height=300)
    
    st.divider()
    
    # Horario cuadrícula
    st.subheader("📅 Horario semanal (vista tradicional)")
    st.caption("Vista hora por hora. Las celdas con — están libres. ⚠️ marca choques reales.")
    
    df_cuadricula, info_choques = construir_horario_cuadricula(horarios)
    
    if df_cuadricula is not None and not df_cuadricula.empty:
        # Calcular altura exacta basada en filas
        num_filas = len(df_cuadricula)
        # Cada fila ocupa aprox 38px + 38px de header
        altura_calculada = 38 + (num_filas * 38) + 3
        
        st.dataframe(
            df_cuadricula,
            use_container_width=True,
            hide_index=True,
            height=altura_calculada
        )
        
        # Mostrar info de choques REALES si existen
        if info_choques:
            st.divider()
            st.warning(f"🚨 **Se detectaron {len(info_choques)} choques REALES en este salón**")
            
            for choque in info_choques:
                c1 = choque['clase_1']
                c2 = choque['clase_2']
                
                clase1_info = c1.get('clases') or {}
                clase2_info = c2.get('clases') or {}
                materia1 = (clase1_info.get('materias') or {}).get('descripcion', '(multi)')
                materia2 = (clase2_info.get('materias') or {}).get('descripcion', '(multi)')
                maestro1 = (clase1_info.get('maestros') or {}).get('nombre_completo', 'Sin asignar')
                maestro2 = (clase2_info.get('maestros') or {}).get('nombre_completo', 'Sin asignar')
                
                with st.expander(f"⚠️ {choque['dia']}: CRN {c1['crn']} vs CRN {c2['crn']}"):
                    col_c1, col_c2 = st.columns(2)
                    with col_c1:
                        st.markdown(f"**CRN {c1['crn']}** (Periodo {c1['periodo_id']})")
                        st.markdown(f"📚 Materia: {materia1}")
                        st.markdown(f"👨‍🏫 Maestro: {maestro1}")
                        st.markdown(f"⏰ Horario: {c1['hora_inicio'][:5]} - {c1['hora_fin'][:5]}")
                    with col_c2:
                        st.markdown(f"**CRN {c2['crn']}** (Periodo {c2['periodo_id']})")
                        st.markdown(f"📚 Materia: {materia2}")
                        st.markdown(f"👨‍🏫 Maestro: {maestro2}")
                        st.markdown(f"⏰ Horario: {c2['hora_inicio'][:5]} - {c2['hora_fin'][:5]}")
                    
                    if maestro1 == maestro2 and maestro1 != 'Sin asignar':
                        st.info("💡 Ambas clases tienen el **mismo maestro** — podría ser una clase espejo.")
                    elif materia1 == materia2:
                        st.info("💡 Ambas son la **misma materia** — podría ser un grupo dividido.")
                    else:
                        st.error("🔴 **Choque real**: dos clases distintas pidiendo el mismo salón al mismo tiempo.")
        
        # Botón de descarga Excel
        st.divider()
        st.subheader("📥 Descargar horario")
        
        excel_bytes = generar_excel_salon(salon_obj, horarios, df_detalle, df_cuadricula)
        
        st.download_button(
            label="📥 Descargar Horario en Excel",
            data=excel_bytes,
            file_name=f"horario_salon_{salon_codigo}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    else:
        st.info("No hay datos suficientes para mostrar la cuadrícula")


main()