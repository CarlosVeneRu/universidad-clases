"""
Página para exportar los datos del sistema a Excel.
Genera un archivo similar al de Banner con los datos actuales.
"""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import io
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.utils.ui import encabezado

from app.utils.queries import get_client, cargar_periodos


encabezado("Exportar Datos", "Descarga la información en formato Excel", "📤")


# Filtros
st.subheader("🎯 Filtros del export")
periodos = cargar_periodos()

def _etq_periodo_exp(p):
    base = f"{p['id']} - {p['descripcion']}"
    if p.get("estado") == "concluido":
        return f"🔒 {base} (Concluido)"
    return base

opciones_periodo = ["Todos los periodos"] + [_etq_periodo_exp(p) for p in periodos]

col_f1, col_f2 = st.columns(2)
with col_f1:
    periodo_sel = st.selectbox("📅 Periodo a exportar", opciones_periodo)
with col_f2:
    modalidad_sel = st.selectbox(
        "🎓 Modalidad",
        ["Todas las modalidades",
         "Solo semestrales (más de 12 semanas)",
         "Solo cuatrimestrales (12 semanas o menos)"]
    )

periodo_filtro = None
if not periodo_sel.startswith("Todos"):
    periodo_filtro = int(periodo_sel.split(" - ")[0])

modalidad_filtro = None
if "semestrales" in modalidad_sel:
    modalidad_filtro = "semestral"
elif "cuatrimestrales" in modalidad_sel:
    modalidad_filtro = "cuatrimestral"

# Toggle para incluir también las clases archivadas en el export
incluir_archivadas = st.toggle(
    "📦 Incluir clases archivadas",
    value=False,
    help="Cuando se activa, el Excel también incluye las clases del histórico (archivadas). "
         "Útil para backups completos o comparativas año a año."
)

st.divider()


def generar_excel(incluir_archivadas=False):
    """Construye el DataFrame con todas las clases y sus horarios.
    Si incluir_archivadas=True, también trae las clases del histórico (clases_archivadas)."""
    client = get_client()

    # 1. Cargar las clases activas (con todos sus datos relacionados)
    query = client.table("clases").select(
        "crn, periodo_id, grupo, clave_periodo, status, "
        "fecha_inicio, fecha_fin, inscritos, capacidad_materia, capacidad_escenario, vacantes, "
        "tipo_curso, tipo_horario, metodo, sin_docente, sin_horario, "
        "impartido_id, impartido_descripcion, modulo_id, modulo_descripcion, "
        "semanas_curso, horas_long, "
        "materias(id, descripcion, grado_materia, area_concentracion, semanas_curso, tipos_uso_compatibles), "
        "maestros(clave, nombre_completo), "
        "carreras(clave_hpn, nombre_hpn, clave_banner, nombre_banner, departamento_id, nivel_id, nivel_descripcion), "
        "campus(id, nombre), "
        "periodos(id, anio, ciclo, descripcion)"
    )

    if periodo_filtro:
        query = query.eq("periodo_id", periodo_filtro)

    # Paginar con ORDER BY estable (evita duplicados y saltos)
    query = query.order("crn").order("periodo_id")

    clases_data = []
    offset = 0
    batch_size = 1000

    while True:
        batch = query.range(offset, offset + batch_size - 1).execute()
        if not batch.data:
            break
        clases_data.extend(batch.data)
        if len(batch.data) < batch_size:
            break
        offset += batch_size

    # Filtro por modalidad (semestral / cuatrimestral) según duración de la clase
    if modalidad_filtro:
        from datetime import date as _date
        def _es_esa_modalidad(c):
            fi, ff = c.get("fecha_inicio"), c.get("fecha_fin")
            if not fi or not ff:
                return False
            try:
                if isinstance(fi, str):
                    fi = _date.fromisoformat(fi)
                if isinstance(ff, str):
                    ff = _date.fromisoformat(ff)
                semanas = (ff - fi).days / 7
            except Exception:
                return False
            if modalidad_filtro == "cuatrimestral":
                return semanas <= 12
            return semanas > 12
        clases_data = [c for c in clases_data if _es_esa_modalidad(c)]

    # Si el toggle está activo, cargar también las clases archivadas
    if incluir_archivadas:
        arch_query = client.table("clases_archivadas").select(
            "crn, periodo_id, grupo, clave_periodo, status, "
            "fecha_inicio, fecha_fin, inscritos, capacidad_materia, capacidad_escenario, vacantes, "
            "tipo_curso, tipo_horario, metodo, sin_docente, sin_horario, "
            "impartido_id, impartido_descripcion, modulo_id, modulo_descripcion, "
            "semanas_curso, horas_long, horarios_snapshot, "
            "materia_id, maestro_clave, carrera_id, campus_id"
        )
        if periodo_filtro:
            arch_query = arch_query.eq("periodo_id", periodo_filtro)
        arch_query = arch_query.order("crn").order("periodo_id")

        arch_data = []
        offset = 0
        while True:
            batch = arch_query.range(offset, offset + 999).execute()
            if not batch.data:
                break
            arch_data.extend(batch.data)
            if len(batch.data) < 1000:
                break
            offset += 1000

        # Aplicar el mismo filtro de modalidad a las archivadas
        if modalidad_filtro:
            from datetime import date as _date
            def _es_esa_modalidad_arch(c):
                fi, ff = c.get("fecha_inicio"), c.get("fecha_fin")
                if not fi or not ff:
                    return False
                try:
                    if isinstance(fi, str):
                        fi = _date.fromisoformat(fi)
                    if isinstance(ff, str):
                        ff = _date.fromisoformat(ff)
                    semanas = (ff - fi).days / 7
                except Exception:
                    return False
                if modalidad_filtro == "cuatrimestral":
                    return semanas <= 12
                return semanas > 12
            arch_data = [c for c in arch_data if _es_esa_modalidad_arch(c)]

        # Resolver los JOINs en memoria (clases_archivadas no tiene FKs directas)
        if arch_data:
            materia_ids = list({c["materia_id"] for c in arch_data if c.get("materia_id")})
            maestro_claves = list({c["maestro_clave"] for c in arch_data if c.get("maestro_clave")})
            carrera_ids = list({c["carrera_id"] for c in arch_data if c.get("carrera_id")})
            campus_ids = list({c["campus_id"] for c in arch_data if c.get("campus_id")})
            periodo_ids = list({c["periodo_id"] for c in arch_data if c.get("periodo_id")})

            mat_map = {}
            if materia_ids:
                mat_map = {m["id"]: m for m in (client.table("materias")
                    .select("id, descripcion, grado_materia, area_concentracion, "
                            "semanas_curso, tipos_uso_compatibles")
                    .in_("id", materia_ids).execute().data or [])}
            maes_map = {}
            if maestro_claves:
                maes_map = {m["clave"]: m for m in (client.table("maestros")
                    .select("clave, nombre_completo")
                    .in_("clave", maestro_claves).execute().data or [])}
            car_map = {}
            if carrera_ids:
                car_map = {c["id"]: c for c in (client.table("carreras")
                    .select("id, clave_hpn, nombre_hpn, clave_banner, nombre_banner, "
                            "departamento_id, nivel_id, nivel_descripcion")
                    .in_("id", carrera_ids).execute().data or [])}
            cam_map = {}
            if campus_ids:
                cam_map = {c["id"]: c for c in (client.table("campus")
                    .select("id, nombre").in_("id", campus_ids).execute().data or [])}
            per_map = {}
            if periodo_ids:
                per_map = {p["id"]: p for p in (client.table("periodos")
                    .select("id, anio, ciclo, descripcion")
                    .in_("id", periodo_ids).execute().data or [])}

            for c in arch_data:
                c["materias"] = mat_map.get(c.get("materia_id"))
                c["maestros"] = maes_map.get(c.get("maestro_clave"))
                c["carreras"] = car_map.get(c.get("carrera_id"))
                c["campus"] = cam_map.get(c.get("campus_id"))
                c["periodos"] = per_map.get(c.get("periodo_id"))
                c["_archivada"] = True  # marca para diferenciar

            clases_data.extend(arch_data)

    if not clases_data:
        return None

    # 2. Cargar TODOS los horarios de las clases activas filtradas
    h_query = client.table("horarios").select(
        "crn, periodo_id, dia_semana, hora_inicio, hora_fin, salon_codigo, es_virtual"
    )
    if periodo_filtro:
        h_query = h_query.eq("periodo_id", periodo_filtro)

    # Paginar con ORDER BY estable
    h_query = h_query.order("crn").order("periodo_id").order("dia_semana")

    horarios_data = []
    offset = 0
    while True:
        batch = h_query.range(offset, offset + 999).execute()
        if not batch.data:
            break
        horarios_data.extend(batch.data)
        if len(batch.data) < 1000:
            break
        offset += 1000

    # Si se incluyen archivadas, expandir los snapshots JSON al mismo formato de horarios
    if incluir_archivadas:
        for c in clases_data:
            if not c.get("_archivada"):
                continue
            snap = c.get("horarios_snapshot") or []
            if not isinstance(snap, list):
                continue
            for h_snap in snap:
                horarios_data.append({
                    "crn": c.get("crn"),
                    "periodo_id": c.get("periodo_id"),
                    "dia_semana": h_snap.get("dia_semana"),
                    "hora_inicio": h_snap.get("hora_inicio"),
                    "hora_fin": h_snap.get("hora_fin"),
                    "salon_codigo": h_snap.get("salon_codigo"),
                    "es_virtual": bool(h_snap.get("es_virtual")),
                })

    # Indexar horarios por (crn, periodo) y día
    horarios_map = {}
    for h in horarios_data:
        clave = (h['crn'], h['periodo_id'])
        if clave not in horarios_map:
            horarios_map[clave] = {}
        horarios_map[clave][h['dia_semana']] = h

    # 3. Construir las filas del DataFrame
    filas = []
    for c in clases_data:
        materia = c.get("materias") or {}
        maestro = c.get("maestros") or {}
        carrera = c.get("carreras") or {}
        campus = c.get("campus") or {}
        periodo = c.get("periodos") or {}

        # Construir horarios por día
        horarios_clase = horarios_map.get((c['crn'], c['periodo_id']), {})

        def get_horario_dia(dia):
            h = horarios_clase.get(dia)
            if not h:
                return "", ""
            hora = f"{str(h['hora_inicio'])[:5]} - {str(h['hora_fin'])[:5]}"
            salon = "VIRTUAL" if h.get('es_virtual') else (h.get('salon_codigo') or "")
            return hora, salon

        h_lun, s_lun = get_horario_dia("LUNES")
        h_mar, s_mar = get_horario_dia("MARTES")
        h_mie, s_mie = get_horario_dia("MIERCOLES")
        h_jue, s_jue = get_horario_dia("JUEVES")
        h_vie, s_vie = get_horario_dia("VIERNES")
        h_sab, s_sab = get_horario_dia("SABADO")
        h_dom, s_dom = get_horario_dia("DOMINGO")

        filas.append({
            "Periodo Banner ID": c.get("periodo_id"),
            "Año ID": periodo.get("anio") if periodo else "",
            "Ciclo ID": periodo.get("ciclo") if periodo else "",
            "Clave Periodo ID": c.get("clave_periodo") or "",
            "Crn ID": c.get("crn"),
            "Grupo ID": c.get("grupo") or "",
            "Materia ID": materia.get("id") if materia else "",
            "Materia DESC": materia.get("descripcion") if materia else "",
            "Grado materia": materia.get("grado_materia") if materia else "",
            "A. concentracion": materia.get("area_concentracion") if materia else "",
            "Semanas del curso": c.get("semanas_curso") or (materia.get("semanas_curso") if materia else "") or "",
            "Tipo uso materia": materia.get("tipos_uso_compatibles") if materia else "",
            "Docente ID": maestro.get("clave") if maestro else "",
            "Docente DESC": maestro.get("nombre_completo") if maestro else "",
            "Carrera Hpn ID": carrera.get("clave_hpn") if carrera else "",
            "Carrera Hpn DESC": carrera.get("nombre_hpn") if carrera else "",
            "Carrera Banner ID": carrera.get("clave_banner") if carrera else "",
            "Carrera Banner DESC": carrera.get("nombre_banner") if carrera else "",
            "Departamento ID": carrera.get("departamento_id") if carrera else "",
            "Nivel ID": carrera.get("nivel_id") if carrera else "",
            "Nivel DESC": carrera.get("nivel_descripcion") if carrera else "",
            "Campus ID": campus.get("id") if campus else "",
            "Fecha Inicio ID": c.get("fecha_inicio") or "",
            "Fecha Final ID": c.get("fecha_fin") or "",
            "Capacidad": c.get("capacidad_materia") or 0,
            "Capacidad Escenario": c.get("capacidad_escenario") or 0,
            "Inscritos": c.get("inscritos") or 0,
            "Vacantes": c.get("vacantes") or 0,
            "Tipo Curso ID": c.get("tipo_curso") or "",
            "Tipo Horario ID": c.get("tipo_horario") or "",
            "Método ID": c.get("metodo") or "",
            "Status ID": c.get("status") or "",
            "Impartición ID": c.get("impartido_id") or "",
            "Desc Impartición ID": c.get("impartido_descripcion") or "",
            "Modulo ID": c.get("modulo_id") or "",
            "desc modulo Ptms": c.get("modulo_descripcion") or "",
            "Sin Docente ID": 1 if c.get("sin_docente") else 0,
            "Sin Horario ID": 1 if c.get("sin_horario") else 0,
            "Horas Long": c.get("horas_long") or "",
            "Horario Lunes ID": h_lun,
            "Salón Lunes ID": s_lun,
            "Horario Martes ID": h_mar,
            "Salón Martes ID": s_mar,
            "Horario Miercoles ID": h_mie,
            "Salón Miercoles ID": s_mie,
            "Horario Jueves ID": h_jue,
            "Salon Jueves ID": s_jue,
            "Horario Viernes ID": h_vie,
            "Salón Viernes ID": s_vie,
            "Horario Sábado ID": h_sab,
            "Salón Sabado ID": s_sab,
            "Horario Domingo ID": h_dom,
            "Salón Domingo ID": s_dom,
            "Es Archivada": "SI" if c.get("_archivada") else "NO",
        })

    return pd.DataFrame(filas)


def df_a_excel_bytes(df, periodo_str):
    """Convierte DataFrame a bytes de Excel con formato."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clases"

    # Encabezados con estilo
    header_fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Datos
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    # Auto-ajustar anchos (con límite)
    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(
            max(len(str(df.columns[col_idx-1])) + 2, 15), 35
        )

    # Congelar encabezado
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# Botón de generación
st.subheader("📥 Generar Excel")

st.info(
    "Se generará un archivo Excel con la misma estructura del reporte original de Banner, "
    "pero con los datos actuales del sistema (incluyendo cualquier cambio manual que se haya hecho)."
)

if st.button("🔨 Generar Excel del sistema", type="primary", use_container_width=True):
    with st.spinner("Generando Excel... esto puede tardar unos segundos"):
        df = generar_excel(incluir_archivadas=incluir_archivadas)

        if df is None or df.empty:
            st.warning("⚠️ No hay datos para exportar con los filtros seleccionados.")
        else:
            # Mostrar previsualización
            total = len(df)
            n_arch = int((df["Es Archivada"] == "SI").sum()) if "Es Archivada" in df.columns else 0
            n_activas = total - n_arch
            if incluir_archivadas and n_arch > 0:
                st.success(f"✅ Excel generado con **{total} clases** ({n_activas} activas + {n_arch} archivadas)")
            else:
                st.success(f"✅ Excel generado con **{total} clases**")

            with st.expander("👁️ Previsualizar datos (primeras 20 filas)"):
                st.dataframe(df.head(20), use_container_width=True, hide_index=True)

            # Generar archivo
            excel_bytes = df_a_excel_bytes(df, periodo_sel)

            fecha_hoy = datetime.now().strftime("%Y%m%d")
            periodo_nombre = "todos" if not periodo_filtro else str(periodo_filtro)
            sufijo_arch = "_con_archivadas" if incluir_archivadas else ""
            nombre_archivo = f"Detalle_Mega_Figpos_Sistema_{periodo_nombre}{sufijo_arch}_{fecha_hoy}.xlsx"

            st.download_button(
                label=f"📥 Descargar {nombre_archivo}",
                data=excel_bytes,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

            st.caption(
                f"💡 Este Excel tiene el mismo formato que el reporte de Banner. "
                f"Puedes guardarlo como respaldo o enviarlo al departamento correspondiente."
            )