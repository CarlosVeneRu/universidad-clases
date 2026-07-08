"""
Funciones compartidas para construir y exportar horarios semanales
estilo cuadrícula. Se usa tanto en Salones como en Maestros.
"""
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


DIAS_ORDEN = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']
DIAS_CORTO = {
    'LUNES': 'LUN', 'MARTES': 'MAR', 'MIERCOLES': 'MIÉ',
    'JUEVES': 'JUE', 'VIERNES': 'VIE', 'SABADO': 'SÁB', 'DOMINGO': 'DOM'
}


def hora_a_minutos(hora_str):
    """Convierte '07:30:00' o '07:30' a minutos desde 00:00."""
    partes = str(hora_str).split(':')
    return int(partes[0]) * 60 + int(partes[1])


def minutos_a_hora(minutos):
    """Convierte minutos a string 'HH:MM'."""
    return f"{minutos // 60:02d}:{minutos % 60:02d}"


def clases_se_solapan(h1, h2):
    """Devuelve True solo si dos clases REALMENTE se solapan en tiempo Y en fechas.
    Considera fecha_inicio / fecha_fin: si una termina antes de que empiece la otra,
    no hay choque real aunque compartan salón y hora."""
    # 1) Verificar solape de horas
    ini1 = hora_a_minutos(h1['hora_inicio'])
    fin1 = hora_a_minutos(h1['hora_fin'])
    ini2 = hora_a_minutos(h2['hora_inicio'])
    fin2 = hora_a_minutos(h2['hora_fin'])
    if not (ini1 < fin2 and ini2 < fin1):
        return False

    # 2) Verificar solape de fechas (si están disponibles)
    def _fechas(h):
        c = h.get('clases') or {}
        return c.get('fecha_inicio'), c.get('fecha_fin')

    fi1, ff1 = _fechas(h1)
    fi2, ff2 = _fechas(h2)
    if fi1 and ff1 and fi2 and ff2:
        # Si una acaba antes de que empiece la otra: no hay choque real
        if ff1 < fi2 or ff2 < fi1:
            return False

    return True


def construir_horario_cuadricula(horarios, etiqueta_extra="salon"):
    """
    Construye horario tipo cuadrícula adaptativo (60 o 30 min).
    
    Args:
        horarios: lista de dicts con dia_semana, hora_inicio, hora_fin, salon_codigo, etc.
        etiqueta_extra: qué mostrar bajo el nombre de la clase:
            - "salon": muestra el salón (uso típico de página Maestros)
            - "materia": muestra la materia (uso típico de página Salones)
    
    Returns:
        (df_cuadricula, info_choques)
    """
    if not horarios:
        return None, None
    
    # Detectar si hay horarios "rotos"
    necesita_media_hora = False
    for h in horarios:
        ini_min = hora_a_minutos(h['hora_inicio'])
        fin_min = hora_a_minutos(h['hora_fin'])
        if ini_min % 60 != 0 or (fin_min + 1) % 60 != 0:
            necesita_media_hora = True
            break
    
    tamano_bloque = 30 if necesita_media_hora else 60
    
    hora_min_minutos = min(hora_a_minutos(h['hora_inicio']) for h in horarios)
    hora_max_minutos = max(hora_a_minutos(h['hora_fin']) for h in horarios)
    
    hora_inicio_grid = (hora_min_minutos // tamano_bloque) * tamano_bloque
    hora_fin_grid = ((hora_max_minutos + tamano_bloque - 1) // tamano_bloque) * tamano_bloque
    
    filas_df = []
    info_choques = []
    
    for bloque_actual in range(hora_inicio_grid, hora_fin_grid, tamano_bloque):
        bloque_inicio_str = minutos_a_hora(bloque_actual)
        bloque_fin_str = minutos_a_hora(bloque_actual + tamano_bloque - 1)
        rango_str = f"{bloque_inicio_str} - {bloque_fin_str}"
        
        fila = {"HORA": rango_str}
        
        for dia in DIAS_ORDEN:
            dia_corto = DIAS_CORTO[dia]
            
            clases_en_bloque = []
            for h in horarios:
                if h['dia_semana'] != dia:
                    continue
                
                clase_ini = hora_a_minutos(h['hora_inicio'])
                clase_fin = hora_a_minutos(h['hora_fin'])
                
                if clase_fin > bloque_actual and clase_ini < bloque_actual + tamano_bloque:
                    clases_en_bloque.append(h)
            
            if not clases_en_bloque:
                fila[dia_corto] = "—"
                continue
            
            hay_solapamiento = False
            if len(clases_en_bloque) > 1:
                # Set para no repetir el mismo par en cada bloque de la cuadrícula
                if 'pares_ya_vistos' not in locals():
                    pares_ya_vistos = set()
                for i in range(len(clases_en_bloque)):
                    for j in range(i+1, len(clases_en_bloque)):
                        if clases_se_solapan(clases_en_bloque[i], clases_en_bloque[j]):
                            hay_solapamiento = True
                            c1 = clases_en_bloque[i]
                            c2 = clases_en_bloque[j]
                            clave_par = tuple(sorted([
                                (c1.get('crn'), c1.get('periodo_id')),
                                (c2.get('crn'), c2.get('periodo_id'))
                            ]))
                            if clave_par not in pares_ya_vistos:
                                pares_ya_vistos.add(clave_par)
                                info_choques.append({
                                    "dia": dia,
                                    "clase_1": c1,
                                    "clase_2": c2
                                })
                            break
                    if hay_solapamiento:
                        break
            
            # Construir texto según el contexto
            textos = []
            for c in clases_en_bloque:
                if etiqueta_extra == "salon":
                    # Contexto Maestros: mostrar Materia + Salón
                    materia = c.get('materia_nombre') or '(sin materia)'
                    salon = c.get('salon_codigo') or ('🌐 Virtual' if c.get('es_virtual') else 'S/Salón')
                    textos.append(f"{materia[:22]} · {salon}")
                else:
                    # Contexto Salones: mostrar Materia + CRN
                    clase_info = c.get('clases') or {}
                    materia_info = (clase_info.get('materias') or {})
                    nombre = materia_info.get('descripcion') or '(multi)'
                    crn = c.get('crn', '')
                    textos.append(f"{nombre[:25]} (CRN {crn})")
            
            if hay_solapamiento:
                fila[dia_corto] = "⚠️ " + " ║ ".join(textos)
            else:
                fila[dia_corto] = " · ".join(textos) if len(textos) > 1 else textos[0]
        
        filas_df.append(fila)
    
    return pd.DataFrame(filas_df), info_choques


def generar_excel_horario(titulo, subtitulo, info_dict, df_detalle, df_cuadricula):
    """
    Genera un Excel con 3 hojas: Resumen, Horario semanal y Detalle.
    
    Args:
        titulo: ej. "HORARIO DEL MAESTRO" o "HORARIO DEL SALÓN"
        subtitulo: ej. el nombre del maestro o código del salón
        info_dict: dict con datos a mostrar en la hoja Resumen ({etiqueta: valor})
        df_detalle: DataFrame con el listado detallado
        df_cuadricula: DataFrame con la cuadrícula
    
    Returns:
        bytes del archivo Excel listo para descargar
    """
    wb = Workbook()
    
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
    bold_font = Font(bold=True)
    
    # ===== HOJA 1: Resumen =====
    ws1 = wb.active
    ws1.title = "Resumen"
    
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"{titulo} · {subtitulo}"
    ws1['A1'].font = titulo_font
    ws1['A1'].fill = titulo_fill
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 25
    
    row = 3
    for etiqueta, valor in info_dict.items():
        ws1.cell(row=row, column=1, value=f"{etiqueta}:").font = bold_font
        ws1.cell(row=row, column=2, value=str(valor))
        row += 1
    
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 40
    
    # ===== HOJA 2: Horario tradicional =====
    ws2 = wb.create_sheet("Horario semanal")
    
    if df_cuadricula is not None and not df_cuadricula.empty:
        ws2.merge_cells('A1:H1')
        ws2['A1'] = f"HORARIO SEMANAL · {subtitulo}"
        ws2['A1'].font = titulo_font
        ws2['A1'].fill = titulo_fill
        ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws2.row_dimensions[1].height = 25
        
        headers = ['HORA', 'LUNES', 'MARTES', 'MIÉRCOLES', 'JUEVES', 'VIERNES', 'SÁBADO', 'DOMINGO']
        header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws2.cell(row=3, column=col_idx, value=header)
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        ocupada_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        libre_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        choque_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        
        for row_idx, row_data in enumerate(df_cuadricula.values, 4):
            for col_idx, valor in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=str(valor))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                cell.font = Font(size=9)
                
                if col_idx == 1:
                    cell.font = Font(bold=True, size=10)
                    cell.fill = header_fill
                elif valor == "—":
                    cell.fill = libre_fill
                    cell.font = Font(size=10, color="9E9E9E")
                elif "⚠️" in str(valor):
                    cell.fill = choque_fill
                else:
                    cell.fill = ocupada_fill
            
            ws2.row_dimensions[row_idx].height = 35
        
        ws2.column_dimensions['A'].width = 15
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws2.column_dimensions[col_letter].width = 25
    
    # ===== HOJA 3: Detalle =====
    ws3 = wb.create_sheet("Detalle")
    
    if df_detalle is not None and not df_detalle.empty:
        for col_idx, col_name in enumerate(df_detalle.columns, 1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1A5490", end_color="1A5490", fill_type="solid")
        
        for row_idx, row_data in enumerate(df_detalle.values, 2):
            for col_idx, valor in enumerate(row_data, 1):
                ws3.cell(row=row_idx, column=col_idx, value=str(valor))
        
        for col_idx in range(1, len(df_detalle.columns) + 1):
            max_len = max(len(str(df_detalle.iloc[i, col_idx-1])) for i in range(len(df_detalle)))
            max_len = max(max_len, len(df_detalle.columns[col_idx-1]))
            ws3.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()